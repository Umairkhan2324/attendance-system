"""app/services/excel_service.py — Excel attendance log writer."""

import logging
import os
import threading
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.config import ExcelConfig

logger = logging.getLogger(__name__)

HEADER_COLOR = "1F4E79"
ALT_ROW_COLOR = "D9E1F2"
COLUMNS = ["Sr. No.", "Employee Code", "Employee Name", "Date", "Time", "Status"]
COL_WIDTHS = [8, 18, 25, 15, 15, 12]


class ExcelService:
    def __init__(self, config: ExcelConfig):
        self.file_path = config.file_path
        self.sheet_name = config.sheet_name
        self._lock = threading.Lock()
        self._init_workbook()

    def _init_workbook(self):
        if os.path.exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
            if self.sheet_name not in self.wb.sheetnames:
                self._create_sheet()
            else:
                self.ws = self.wb[self.sheet_name]
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name
            self._write_headers()
            self.wb.save(self.file_path)
            logger.info(f"Created Excel file: {self.file_path}")

    def _create_sheet(self):
        self.ws = self.wb.create_sheet(self.sheet_name)
        self._write_headers()
        self.wb.save(self.file_path)

    def _write_headers(self):
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        for col, header in enumerate(COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, (letter, width) in enumerate(
            zip(["A", "B", "C", "D", "E", "F"], COL_WIDTHS), 1
        ):
            self.ws.column_dimensions[letter].width = width

    def log(
        self,
        employee_code: str,
        employee_name: str,
        status: str = "Present",
    ) -> tuple[str, str]:
        """Write one attendance row. Returns (date_str, time_str)."""
        with self._lock:
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")

            next_row = self.ws.max_row + 1
            sr_no = next_row - 1

            row_data = [sr_no, employee_code, employee_name, date_str, time_str, status]
            alt_fill = PatternFill(
                start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid"
            )

            for col, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=next_row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                if next_row % 2 == 0:
                    cell.fill = alt_fill

            self.wb.save(self.file_path)
            logger.info(
                f"Excel logged — {employee_code} | {employee_name} | {date_str} {time_str}"
            )
            return date_str, time_str

    def get_all_records(self) -> list[dict]:
        """Read all rows from the sheet and return as list of dicts."""
        with self._lock:
            records = []
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row[1]:  # employee_code must exist
                    records.append(
                        {
                            "sr_no": row[0],
                            "employee_code": row[1],
                            "employee_name": row[2],
                            "date": row[3],
                            "time": row[4],
                            "status": row[5],
                        }
                    )
            return records
